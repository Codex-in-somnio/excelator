#!/usr/bin/env python3

from flask import Flask, render_template, request
from urllib.parse import quote_plus, unquote_plus
from flask_sqlalchemy import SQLAlchemy
from waitress import serve

import os
import openpyxl
import re
import threading
import json
import logging
import coloredlogs

import config


DATA_PATH = 'data'
COMPLETE_LIST_JSON_PATH = 'data/completed_list.json'

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///data/backup/backup.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

lock = threading.Lock()
cl_json_lock = threading.Lock()

coloredlogs.install(level=logging.DEBUG)


class DbCell(db.Model):
    filename = db.Column(db.String, primary_key=True)
    worksheet = db.Column(db.String, primary_key=True)
    cell_coord = db.Column(db.String, primary_key=True)
    timestamp = db.Column(db.DateTime, primary_key=True,
                          server_default=db.func.now())
    text = db.Column(db.String, nullable=True)


class Cell:
    coord = None
    text = None
    merged = False
    skip = False
    edit = False
    style = ""
    tab_index = 0
    hidden = False
    char_limit = None

    def __init__(self, cell, merged_cell_ranges):
        self.c = cell

        def check_cell_merged(coord):
            for merged_cell_range in merged_cell_ranges:
                if coord in merged_cell_range:
                    return merged_cell_range
            return False

        merged_cell_range = check_cell_merged(cell.coordinate)
        if merged_cell_range:
            if merged_cell_range.start_cell == cell:
                self.merged = (len(merged_cell_range.top),
                               len(merged_cell_range.left))
            else:
                self.skip = True
                return

        def cell_value_to_text(val):
            if isinstance(val, str) and val[0] == '=':
                m = re.search(r'=LEN\(([A-Za-z0-9]+)\)', val)
                if m:
                    c = m.group(1)
                    len_cell_val = cell.parent[c].value
                    if isinstance(len_cell_val, str):
                        return str(len(cell.parent[c].value))
                    else:
                        return '#VAL_ERR'
                else:
                    '#FUNC_NOT_IMPLEMENTED'
            else:
                return str(val) if val else ''

        self.text = cell_value_to_text(cell.value)

        self.coord = cell.coordinate

        styles = {
            'overflow': 'hidden'
        }

        def get_color(color):
            if color.type == 'indexed':
                index = color.indexed - 1
                return '#' + openpyxl.styles.colors.COLOR_INDEX[index]
            elif color.type == 'rgb':
                return '#' + color.rgb
            return None

        # fg_color = get_color(cell.fill.fgColor)
        # if fg_color:
        #     styles['color'] = fg_color

        # bg_color = get_color(cell.fill.bgColor)
        # if bg_color:
        #     styles['background-color'] = bg_color

        for attr, value in styles.items():
            self.style += f'{attr}:{value};'

        if cell.column_letter in config.EDIT_COLS and \
                cell.row > config.HEADER_ROWS:
            self.edit = True
            self.tab_index = 100 + \
                cell.row + (cell.column - 1) * cell.parent.max_row

        if cell.column_letter in config.HIDE_COLS:
            self.hidden = True

        if self.edit and cell.column_letter == config.CHAR_LIMIT_APPLY_COL:
            try:
                char_lim_cell = cell.parent.cell(
                    cell.row,
                    ord(config.CHAR_LIMIT_VAL_COL) - ord('A') + 1
                )
                merged_cell_range = check_cell_merged(char_lim_cell.coordinate)
                if merged_cell_range:
                    char_lim_cell = merged_cell_range.start_cell

                char_lim_val = cell_value_to_text(char_lim_cell.value)

                if config.CHAR_LIMIT_REGEX:
                    result = re.search(config.CHAR_LIMIT_REGEX, char_lim_val)
                    if result and len(result.groups()) == 1:
                        self.char_limit = int(result.group(1))
                    elif not result:
                        logging.debug(
                            'failed to extract char limit from '
                            f'{char_lim_cell.coordinate} for {self.coord}')
                    else:
                        logging.debug(
                            'regex match result should only contain one capture'
                            ' group.')
                else:
                    self.char_limit = int(
                        int(re.sub(r'[^\d]+', '', char_lim_val))
                    )

            except ValueError:
                logging.debug(
                    f'failed to parse int from {char_lim_cell.coordinate} '
                    f'for {self.coord}')


@app.template_filter('urlencode')
def urlencode_filter(s):
    return quote_plus(s)


@app.route('/')
def index():
    search_for = request.args.get('search')

    if search_for:
        logging.debug(f'搜索：{search_for}')

    files = []
    search_result_list = []

    for f in os.listdir(DATA_PATH):
        if os.path.isfile(os.path.join(DATA_PATH, f)) and \
                f.lower().endswith('.xlsx') and \
                not f.startswith('~$'):
            if search_for and search_xlsx_for(f, search_for):
                logging.debug(f'{f}包含搜索关键字')
                search_result_list.append(f)
            files.append(f)
    files.sort()
    cur_file = request.args.get('filename')
    cur_ws = request.args.get('worksheet')

    with open(COMPLETE_LIST_JSON_PATH) as f:
        complete_list = json.load(f)

    if not cur_file:
        return render_template('body.html', files=files, search_for=search_for,
                               complete_list=complete_list,
                               search_result_list=search_result_list)

    wb = openpyxl.open(os.path.join(DATA_PATH, cur_file))

    if cur_ws:
        ws = wb[cur_ws]
    else:
        ws = wb.active
        cur_ws = ws.title

    table = []
    msg = ''

    col_edge = 0
    row_edge = 0

    at_least_col = sorted(config.EDIT_COLS)[-1]
    ws[at_least_col + '1']

    if ws.max_row > 1:
        for row in ws:
            _row = []
            row_empty = True
            for cell in row:
                _cell = Cell(cell, ws.merged_cells.ranges)
                _row.append(_cell)
                if cell.value or _cell.edit:
                    col_edge = max(col_edge, cell.column)
                    row_empty = False
            table.append(_row)
            if not row_empty:
                row_edge = cell.row

        table = table[:row_edge]
        for i in range(len(table)):
            table[i] = table[i][:col_edge]
    else:
        msg = '无内容'

    wss = [ws.title for ws in wb.worksheets]

    rendered = render_template('body.html',
                               files=files, cur_file=cur_file, cur_ws=cur_ws,
                               table=table, wss=wss, msg=msg,
                               search_for=search_for,
                               complete_list=complete_list,
                               search_result_list=search_result_list)
    wb.close()
    return rendered


@app.route('/write', methods=['POST'])
def write():
    data = request.get_json()
    with lock:
        logging.debug(f'收到xlsx写入数据：{data}')
        wb = openpyxl.open(os.path.join(DATA_PATH, data['filename']))
        ws = wb[data['worksheet']]
        for cell, val in data['cells'].items():
            ws[cell].value = val
            db_cell = DbCell(filename=data['filename'],
                             worksheet=data['worksheet'],
                             cell_coord=cell, text=val)
            db.session.add(db_cell)
        db.session.commit()
        wb.save(os.path.join(DATA_PATH, data['filename']))
        wb.close()

    return 'ok'


@app.route('/completed_list', methods=['PUT', 'DELETE'])
def completed_list():
    data = request.get_json()
    method = request.method
    filename = data['filename']
    with cl_json_lock:
        logging.debug(f'收到完成列表请求：{data}, {method}')
        with open(COMPLETE_LIST_JSON_PATH) as f:
            cur_list = json.load(f)
        if method == 'DELETE':
            while filename in cur_list:
                cur_list.remove(filename)
        else:
            if filename not in cur_list:
                cur_list.append(filename)
        with open(COMPLETE_LIST_JSON_PATH, 'w') as f:
            json.dump(cur_list, f, indent=2, ensure_ascii=False)

    return 'ok'


def search_xlsx_for(filename, search_for, editable_only=False):
    '''
    returns: {(sheet, cell_coord): (val, [pos, ...]), ...}
    '''
    search_for = search_for.lower()
    ret = {}
    wb = openpyxl.open(os.path.join(DATA_PATH, filename), read_only=True)
    for ws in wb:
        for row_number, row in enumerate(ws.iter_rows()):
            if row_number < config.HEADER_ROWS:
                continue
            for cell in row:
                if isinstance(cell, openpyxl.cell.read_only.EmptyCell) or \
                        cell.column_letter in config.HIDE_COLS:
                    continue
                if editable_only and cell.column_letter not in config.EDIT_COLS:
                    continue
                cell_val = cell.value
                if cell_val is not None:
                    cell_val = str(cell_val)
                    if cell_val.startswith('='):
                        continue
                    start_pos = 0
                    while True:
                        found_pos = cell_val.lower().find(search_for, start_pos)
                        if found_pos == -1:
                            break
                        
                        logging.debug(f'hit: {cell_val} at pos {start_pos}')
                        
                        ret.setdefault(
                            (ws.title, cell.coordinate), (cell_val, []))
                        ret[(ws.title, cell.coordinate)][1].append(found_pos)

                        start_pos = found_pos + len(search_for)
    wb.close()
    return ret


if __name__ == '__main__':
    if not os.path.isfile('data/backup/backup.db'):
        logging.warning('备份数据库文件不存在，即将创建')
        db.create_all()
    if not os.path.isfile('data/completed_list.json'):
        logging.warning('完成列表文件不存在，即将创建')
        with open(COMPLETE_LIST_JSON_PATH, 'w') as f:
            json.dump([], f)

    serve(app, host='127.0.0.1', port=5001)
    # app.run(debug=True)
